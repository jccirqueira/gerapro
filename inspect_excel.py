import openpyxl

def inspect_excel():
    print("Opening workbook...")
    wb = openpyxl.load_workbook('DVT260006-PTC_00.xlsm', data_only=False)
    
    for sheet_name in wb.sheetnames:
        if 'PC' in sheet_name or 'Custos' in sheet_name or 'DP' in sheet_name:
            print(f"\n=== SHEET: {sheet_name} ===")
            sheet = wb[sheet_name]
            
            for row in sheet.iter_rows():
                row_data = f"Row {row[0].row}: "
                has_data = False
                for cell in row:
                    if cell.data_type == 'f': # Formula
                        has_data = True
                        row_data += f"[Col {cell.column_letter}: FORMULA -> {cell.value}] "
                    elif isinstance(cell.value, str):
                        val_lower = cell.value.lower()
                        if any(kw in val_lower for kw in ['imposto', 'margem', 'total', 'lucro', 'custo', 'venda', 'desconto', 'bdi', 'encargo']):
                            has_data = True
                            row_data += f"[Col {cell.column_letter}: LABEL -> {cell.value}] "
                    elif isinstance(cell.value, (int, float)):
                        row_data += f"[Col {cell.column_letter}: NUM -> {cell.value}] "
                
                if has_data:
                    print(row_data)

if __name__ == "__main__":
    inspect_excel()
